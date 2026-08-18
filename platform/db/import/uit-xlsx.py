#!/usr/bin/env python3
"""
'1. Creative Strategy Map.xlsx'  ->  creative-strategy-map.csv

    pip install openpyxl
    python3 platform/db/import/uit-xlsx.py ~/Downloads/'1. Creative Strategy Map.xlsx'

Het bestand staat in Drive (id 11vHqpihyhh3DYyPsa9dpdT7eBcf8dajU) en niet in
Git: het is 447 kB, het verandert dagelijks, en de veertien tabbladen die dit
script niet leest horen niet in een repository thuis. Wat hier wel in staat is
de uitdraai, zodat te zien is wat er geïmporteerd is.

Twee dingen die niet vanzelf goed gaan
──────────────────────────────────────

1. Lees de xlsx, niet een tekstweergave. De eerste poging ging via de
   platte-tekstuitvoer van Drive, en die sloopt elk accentteken: 'één pakket'
   werd ' n pakket', in 210 velden. Dat valt bij steekproeven niet op omdat de
   zin verder klopt. Tellen wel: nul van de 624 rijen had nog een é, à of ë.

2. Stop bij 'Breakdown Analyses'. Onder de advertenties staat het analyseblok
   van de sheet zelf, en dat heeft ook rijen met tekst in de naamkolom — 'Item',
   'Mark (Pragmatist)', '🎠 Carousel (Meta / Instagram)'. Zonder die grens
   importeer je 45 kopregels als advertentie, en drie ervan heten 'Item'.
   Met de grens komen er 624 rijen uit, precies wat de sheet zelf op zijn
   dashboardkaart 'TOTAL TESTS' zet.
"""

import csv
import datetime
import os
import sys

TAB = '📊 Test Tracker'
KOPRIJ = 11        # de kolomnamen
EERSTE = 12        # de eerste advertentie
NAAMKOLOM = 2      # 'Ad Name'


def main():
    if len(sys.argv) != 2:
        sys.exit(__doc__.strip().splitlines()[2].strip())
    try:
        import openpyxl
    except ImportError:
        sys.exit('openpyxl ontbreekt: pip install openpyxl')

    wb = openpyxl.load_workbook(sys.argv[1], data_only=True)
    if TAB not in wb.sheetnames:
        sys.exit(f'tabblad {TAB!r} ontbreekt; gevonden: {wb.sheetnames}')
    ws = wb[TAB]

    grens = None
    for r in range(EERSTE, ws.max_row + 1):
        if any(ws.cell(r, c).value and 'Breakdown' in str(ws.cell(r, c).value)
               for c in range(1, 6)):
            grens = r
            break
    if grens is None:
        sys.exit("de grens 'Breakdown Analyses' is niet gevonden — is de sheet "
                 "verbouwd? Zonder die grens komt het analyseblok mee als data.")

    kols = [(c, ws.cell(KOPRIJ, c).value)
            for c in range(NAAMKOLOM, ws.max_column + 1)
            if ws.cell(KOPRIJ, c).value]

    uit = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       'creative-strategy-map.csv')
    n = 0
    with open(uit, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['bron_rij'] + [k for _, k in kols])
        for r in range(EERSTE, grens):
            naam = ws.cell(r, NAAMKOLOM).value
            if naam is None or str(naam).strip() == '':
                continue
            rij = []
            for c, _ in kols:
                v = ws.cell(r, c).value
                if isinstance(v, (datetime.datetime, datetime.date)):
                    v = v.strftime('%Y-%m-%d')
                elif v is None:
                    v = ''
                else:
                    v = str(v).strip()
                rij.append(v)
            n += 1
            w.writerow([n] + rij)

    print(f'{n} advertenties -> {uit}', file=sys.stderr)


if __name__ == '__main__':
    main()
